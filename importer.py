"""Bulk import for Collecta: instalments and payments.

Flow: download an Excel template (with an Instructions/mapping sheet) -> paste rows
-> upload -> validate (preview) -> confirm -> audit report with old/new balances.
"""
import io
import csv
from datetime import datetime, date

import services as svc

# ---- column specs: (key, header, required, help, example) ----
SPECS = {
    "instalments": dict(
        title="Instalments",
        columns=[
            ("cust_ref", "Customer Ref", "No", "Leave blank for a new customer and Collecta assigns the next code; or give an existing ref", "(blank) or EGP049"),
            ("customer_name", "Customer Name", "If new", "Required for a new customer", "شركة النور"),
            ("currency", "Currency", "If new", "EGP or USD (must match an existing customer)", "EGP"),
            ("account_no", "Account No", "No", "Optional account number for a new customer", "900123"),
            ("inst_id", "Instalment ID", "Yes", "Unique ID for this instalment", "EGP049-01"),
            ("original_amount", "Original Amount", "Yes", "Instalment amount (number, no currency symbol)", "50000"),
            ("due_date", "Due Date", "Yes", "Date the instalment is due (YYYY-MM-DD)", "2026-03-15"),
            ("date_raised", "Date Raised", "No", "Date the instalment was raised (YYYY-MM-DD)", "2026-01-15"),
            ("security", "Security", "No", "Cheque / Trust Receipt / Unsecured / etc.", "Cheque"),
            ("reference", "Reference", "No", "Cheque / invoice / contract number", "CHQ-88213"),
            ("description", "Description", "No", "Free-text note", "Q1 delivery"),
        ],
    ),
    "payments": dict(
        title="Payments (receipts)",
        columns=[
            ("inst_id", "Instalment ID", "Yes", "The instalment this receipt pays (must already exist)", "EGP049-01"),
            ("amount", "Amount", "Yes", "Amount received (number, no currency symbol)", "20000"),
            ("method", "Method", "Yes", "Cash / Cheque / Bank Transfer / Card / Other", "Bank Transfer"),
            ("collected_on", "Collected On", "Yes", "Actual date received (YYYY-MM-DD)", "2026-05-02"),
            ("txn_ref", "Transaction Ref", "No", "Receipt / transaction reference", "RCPT-1201"),
            ("received_by", "Received By", "No", "Defaults to the person importing", "Ahmed"),
        ],
    ),
}


# ---------------- template ----------------
def build_template(kind):
    """Return xlsx bytes: a Data sheet (headers + 2 example rows) and an Instructions sheet."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    spec = SPECS[kind]
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    headers = [c[1] for c in spec["columns"]]
    navy = PatternFill("solid", fgColor="1F3864")
    bold_white = Font(bold=True, color="FFFFFF")
    for j, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=j, value=h)
        cell.fill = navy; cell.font = bold_white
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[cell.column_letter].width = max(14, len(h) + 3)
    # two example rows (grey) the user overwrites
    grey = Font(color="9AA0A6", italic=True)
    if kind == "instalments":
        ex1 = ["", "شركة النور", "EGP", "900123", "NEW-01", "50000", "2026-03-15", "2026-01-15", "Cheque", "CHQ-88213", "new customer — ref left blank, auto-assigned"]
        ex2 = ["EGP001", "", "", "", "EGP001-99", "25000", "2026-06-15", "", "Trust Receipt", "TR-9", "existing customer — give its ref"]
    else:
        ex1 = [c[4] for c in spec["columns"]]
        ex2 = ["EGP049-01", "10000", "Cash", "2026-06-02", "RCPT-1202", ""]
    for r, ex in enumerate((ex1, ex2), start=2):
        for j, v in enumerate(ex, 1):
            c = ws.cell(row=r, column=j, value=v); c.font = grey
    ws.freeze_panes = "A2"
    # instructions sheet
    ins = wb.create_sheet("Instructions")
    ins.append(["Column", "Required?", "What to enter", "Example"])
    for j in range(1, 5):
        c = ins.cell(row=1, column=j); c.fill = navy; c.font = bold_white
    for key, header, req, help_, ex in spec["columns"]:
        ins.append([header, req, help_, ex])
    for col, w in zip("ABCD", (22, 12, 52, 22)):
        ins.column_dimensions[col].width = w
    ins.insert_rows(1)
    ins.cell(row=1, column=1, value=f"Collecta bulk import — {spec['title']}. "
             f"Paste your rows under the headers on the 'Data' sheet, in the same order. "
             f"Delete the two grey example rows before uploading.")
    ins.cell(row=1, column=1).font = Font(bold=True, color="1F3864")
    buf = io.BytesIO(); wb.save(buf); return buf.getvalue()


# ---------------- parsing ----------------
def _norm(s):
    return "".join(ch for ch in str(s or "").strip().lower() if ch.isalnum())


def read_table(file_storage, kind):
    """Read an uploaded xlsx or csv into a list of dict rows keyed by spec keys."""
    spec = SPECS[kind]
    header_to_key = {_norm(h): k for k, h, *_ in spec["columns"]}
    # also allow the internal key itself as a header
    for k, h, *_ in spec["columns"]:
        header_to_key.setdefault(_norm(k), k)
    name = (file_storage.filename or "").lower()
    rows = []
    raw = file_storage.read()
    if name.endswith(".csv") or (not name.endswith((".xlsx", ".xlsm"))):
        text = raw.decode("utf-8-sig", errors="replace")
        reader = csv.reader(io.StringIO(text))
        table = list(reader)
    else:
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(raw), data_only=True, read_only=True)
        ws = wb["Data"] if "Data" in wb.sheetnames else wb[wb.sheetnames[0]]
        table = [[c for c in row] for row in ws.iter_rows(values_only=True)]
    if not table:
        return []
    header_row = table[0]
    idx = {}
    for j, h in enumerate(header_row):
        k = header_to_key.get(_norm(h))
        if k:
            idx[k] = j
    for r_i, row in enumerate(table[1:], start=2):
        if row is None or all((c is None or str(c).strip() == "") for c in row):
            continue
        rec = {"_row": r_i}
        for k, j in idx.items():
            v = row[j] if j < len(row) else None
            rec[k] = v
        rows.append(rec)
    return rows


def _to_date(v):
    if v is None or str(v).strip() == "":
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    s = str(v).strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d-%m-%Y", "%Y/%m/%d", "%d %b %Y", "%d-%b-%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return "ERR"


def _to_amount(v):
    if v is None or str(v).strip() == "":
        return None
    try:
        return float(str(v).replace(",", "").replace(" ", ""))
    except ValueError:
        return "ERR"


# ---------------- validation ----------------
def validate(kind, rows):
    """Return (accepted, rejected). accepted rows are normalized dicts ready to commit."""
    from models import Customer, Instalment
    accepted, rejected = [], []
    existing_cust = {c.cust_ref.upper(): c for c in Customer.query.all() if c.cust_ref}
    existing_inst = {i.inst_id.upper() for i in Instalment.query.all() if i.inst_id}
    seen_inst = set()

    if kind == "instalments":
        for r in rows:
            errs = []
            ref = str(r.get("cust_ref") or "").strip()
            inst_id = str(r.get("inst_id") or "").strip()
            amt = _to_amount(r.get("original_amount"))
            due = _to_date(r.get("due_date"))
            raised = _to_date(r.get("date_raised"))
            ccy = str(r.get("currency") or "").strip().upper()
            name = str(r.get("customer_name") or "").strip()
            if not inst_id: errs.append("Instalment ID is required")
            if amt in (None, "ERR") or (isinstance(amt, float) and amt <= 0):
                errs.append("Original Amount must be a positive number")
            if due in (None, "ERR"): errs.append("Due Date is missing or unreadable")
            if raised == "ERR": errs.append("Date Raised is unreadable")
            if ref:
                cust = existing_cust.get(ref.upper())
                is_new_customer = cust is None
                if is_new_customer:
                    if not name: errs.append("Customer Name is required for a new customer")
                    if ccy not in ("EGP", "USD"): errs.append("Currency (EGP/USD) is required for a new customer")
                else:
                    if ccy and ccy != cust.currency:
                        errs.append(f"Currency {ccy} does not match existing customer ({cust.currency})")
                    ccy = ccy or cust.currency
            else:
                # blank ref -> a new customer; Collecta will auto-assign the next code
                is_new_customer = True
                if not name: errs.append("Provide a Customer Ref, or a Customer Name for a new customer")
                if ccy not in ("EGP", "USD"): errs.append("Currency (EGP/USD) is required for a new customer")
            if inst_id and inst_id.upper() in existing_inst:
                errs.append("Instalment ID already exists in the system")
            if inst_id and inst_id.upper() in seen_inst:
                errs.append("Instalment ID is duplicated within the file")
            if inst_id: seen_inst.add(inst_id.upper())
            if errs:
                rejected.append(dict(row=r["_row"], key=inst_id or ref or name, reason="; ".join(errs)))
            else:
                accepted.append(dict(_row=r["_row"], cust_ref=ref, customer_name=name,
                                     currency=ccy,
                                     account_no=str(r.get("account_no") or "").strip() or None,
                                     inst_id=inst_id, original_amount=amt,
                                     due_date=due.isoformat() if due else None,
                                     date_raised=raised.isoformat() if isinstance(raised, date) else None,
                                     security=str(r.get("security") or "").strip() or None,
                                     reference=str(r.get("reference") or "").strip() or None,
                                     description=str(r.get("description") or "").strip() or None,
                                     is_new_customer=is_new_customer,
                                     auto_ref=(is_new_customer and not ref)))
    else:  # payments
        inst_by_id = {i.inst_id.upper(): i for i in Instalment.query.all() if i.inst_id}
        cust_by_id = {c.id: c for c in Customer.query.all()}
        for r in rows:
            errs = []
            inst_id = str(r.get("inst_id") or "").strip()
            amt = _to_amount(r.get("amount"))
            method = str(r.get("method") or "").strip()
            on = _to_date(r.get("collected_on"))
            if not inst_id: errs.append("Instalment ID is required")
            elif inst_id.upper() not in inst_by_id: errs.append("Instalment ID not found in the system")
            if amt in (None, "ERR") or (isinstance(amt, float) and amt <= 0):
                errs.append("Amount must be a positive number")
            if not method: errs.append("Method is required")
            if on in (None, "ERR"): errs.append("Collected On is missing or unreadable")
            if errs:
                rejected.append(dict(row=r["_row"], key=inst_id, reason="; ".join(errs)))
            else:
                inst = inst_by_id[inst_id.upper()]
                cust = cust_by_id.get(inst.customer_id)
                accepted.append(dict(_row=r["_row"], inst_id=inst.inst_id, amount=amt,
                                     method=method, collected_on=on.isoformat(),
                                     txn_ref=str(r.get("txn_ref") or "").strip() or None,
                                     received_by=str(r.get("received_by") or "").strip() or None,
                                     currency=inst.currency,
                                     cust_ref=(cust.cust_ref if cust else ""),
                                     _ref=(cust.cust_ref.upper() if cust and cust.cust_ref else "")))
    return accepted, rejected


# ---------------- balances snapshot ----------------
def balances_by_customer():
    """cust_ref -> dict(name, currency, net)."""
    out = {}
    for b in svc.customer_balances():
        out[b["cust_ref"].upper()] = dict(name=b["customer"], currency=b["currency"], net=b["total"])
    return out


# ---------------- commit ----------------
def commit(kind, accepted, user):
    from models import db, Customer, Instalment, Collection
    created = 0
    affected_refs = set()
    new_customers = []
    if kind == "instalments":
        cust_by_ref = {c.cust_ref.upper(): c for c in Customer.query.all() if c.cust_ref}
        reserved = set(cust_by_ref.keys())
        group_ref = {}   # (normalized name, currency) -> auto-assigned ref, for blank-ref rows

        def resolve_ref(a):
            ref = (a.get("cust_ref") or "").strip()
            if ref:
                return ref
            key = (_norm(a.get("customer_name")), a.get("currency"))
            if key not in group_ref:
                nr = svc.next_customer_ref(a.get("currency"), reserved)
                reserved.add(nr.upper())
                group_ref[key] = nr
            return group_ref[key]

        for a in accepted:
            ref = resolve_ref(a)
            cust = cust_by_ref.get(ref.upper())
            if cust is None:
                cust = Customer(cust_ref=ref, name=a["customer_name"], currency=a["currency"],
                                account_no=a.get("account_no"),
                                owner_id=svc.default_owner_for(a["currency"]))
                db.session.add(cust); db.session.flush()
                cust_by_ref[ref.upper()] = cust
                new_customers.append(ref)
            inst = Instalment(inst_id=a["inst_id"], customer_id=cust.id, currency=cust.currency,
                              original_amount=a["original_amount"],
                              due_date=date.fromisoformat(a["due_date"]) if a.get("due_date") else None,
                              date_raised=date.fromisoformat(a["date_raised"]) if a.get("date_raised") else None,
                              security=a.get("security"), reference=a.get("reference"),
                              description=a.get("description"))
            db.session.add(inst); created += 1
            affected_refs.add(ref.upper())
        db.session.commit()
    else:
        inst_by_id = {i.inst_id.upper(): i for i in Instalment.query.all() if i.inst_id}
        cust_ref_by_id = {c.id: c.cust_ref for c in Customer.query.all()}
        for a in accepted:
            inst = inst_by_id.get(a["inst_id"].upper())
            if not inst:
                continue
            on = date.fromisoformat(a["collected_on"])
            dov = (on - inst.due_date).days if inst.due_date else None
            bucket = svc.bucket_for(dov)
            col = Collection(instalment_id=inst.id, customer_id=inst.customer_id,
                             currency=inst.currency, amount=a["amount"], method=a["method"],
                             collected_on=on, txn_ref=a.get("txn_ref"),
                             received_by=a.get("received_by") or user,
                             bucket_at_collection=bucket)
            db.session.add(col); created += 1
            ref = cust_ref_by_id.get(inst.customer_id)
            if ref: affected_refs.add(ref.upper())
        db.session.commit()
    return dict(created=created, affected_refs=affected_refs, new_customers=new_customers)


# ---------------- audit xlsx ----------------
def build_audit(kind, meta, accepted, rejected, before, after):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    navy = PatternFill("solid", fgColor="1F3864"); bw = Font(bold=True, color="FFFFFF")
    wb = Workbook(); s = wb.active; s.title = "Summary"
    s["A1"] = f"Collecta bulk-import audit — {SPECS[kind]['title']}"; s["A1"].font = Font(bold=True, size=14, color="1F3864")
    rows = [("Imported by", meta.get("user")), ("Date/time", meta.get("when")),
            ("Source file", meta.get("filename")), ("Rows accepted", len(accepted)),
            ("Rows rejected", len(rejected))]
    for i, (k, v) in enumerate(rows, start=3):
        s.cell(row=i, column=1, value=k).font = Font(bold=True); s.cell(row=i, column=2, value=v)
    s.column_dimensions["A"].width = 20; s.column_dimensions["B"].width = 46

    # balances sheet
    bs = wb.create_sheet("Balance movement")
    hdr = ["Customer Ref", "Customer", "Currency", "Old balance", "Change", "New balance"]
    bs.append(hdr)
    for j in range(1, len(hdr) + 1):
        bs.cell(row=1, column=j).fill = navy; bs.cell(row=1, column=j).font = bw
    tot = {}
    for ref in sorted(meta.get("affected_refs", [])):
        b = before.get(ref, {}); a = after.get(ref, {})
        ccy = a.get("currency") or b.get("currency") or ""
        old = b.get("net", 0.0); new = a.get("net", 0.0)
        bs.append([ref, a.get("name") or b.get("name") or "", ccy, round(old, 2),
                   round(new - old, 2), round(new, 2)])
        t = tot.setdefault(ccy, [0.0, 0.0]); t[0] += old; t[1] += new
    for w, col in zip((14, 34, 10, 16, 14, 16), "ABCDEF"):
        bs.column_dimensions[col].width = w
    bs.append([])
    for ccy, (old, new) in tot.items():
        row = ["TOTAL", "", ccy, round(old, 2), round(new - old, 2), round(new, 2)]
        bs.append(row)
        for j in range(1, 7):
            bs.cell(row=bs.max_row, column=j).font = Font(bold=True)

    # accepted sheet
    aa = wb.create_sheet("Accepted")
    cols = [c[1] for c in SPECS[kind]["columns"]]
    keys = [c[0] for c in SPECS[kind]["columns"]]
    aa.append(["File row"] + cols)
    for j in range(1, len(cols) + 2):
        aa.cell(row=1, column=j).fill = navy; aa.cell(row=1, column=j).font = bw
    for a in accepted:
        aa.append([a.get("_row")] + [a.get(k) for k in keys])
    # rejected sheet
    rj = wb.create_sheet("Rejected")
    rj.append(["File row", "Key", "Reason"])
    for j in range(1, 4):
        rj.cell(row=1, column=j).fill = navy; rj.cell(row=1, column=j).font = bw
    for r in rejected:
        rj.append([r["row"], r["key"], r["reason"]])
    rj.column_dimensions["C"].width = 70; rj.column_dimensions["B"].width = 18
    buf = io.BytesIO(); wb.save(buf); return buf.getvalue()


def projected(kind, accepted, before):
    """For the preview: per-customer old -> projected new balance (and totals)."""
    move = {}
    for a in accepted:
        if kind == "instalments":
            ref = (a.get("cust_ref") or "").strip()
            if ref:
                key = ref.upper(); disp = ref
            else:  # blank ref -> new customer grouped by name; ref auto-assigned at commit
                key = "NEW:" + _norm(a.get("customer_name")) + ":" + (a.get("currency") or "")
                disp = "(auto)"
        else:
            key = a.get("_ref")
            disp = a.get("cust_ref", key)
        if not key:
            continue
        m = move.setdefault(key, dict(ref=disp, name=a.get("customer_name", "") if kind == "instalments" else "",
                                      currency=a.get("currency", ""), old=0.0, change=0.0))
        if kind == "instalments":
            m["change"] += float(a["original_amount"])
            if a.get("customer_name"): m["name"] = a["customer_name"]
        else:
            m["change"] -= float(a["amount"])
    rows, tot = [], {}
    for key, m in move.items():
        b = before.get(key, {})
        old = b.get("net", 0.0)
        name = b.get("name") or m["name"] or ""
        ccy = b.get("currency") or m["currency"] or ""
        new = old + m["change"]
        rows.append(dict(ref=m["ref"], name=name, currency=ccy, old=old, change=m["change"], new=new))
        t = tot.setdefault(ccy, dict(old=0.0, change=0.0, new=0.0))
        t["old"] += old; t["change"] += m["change"]; t["new"] += new
    rows.sort(key=lambda x: x["ref"])
    return rows, tot

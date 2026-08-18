"""Reusable report exporters: Excel (.xlsx) and PDF, with the organisation letterhead.
PDF renders Arabic correctly (embedded Amiri font + reshaping/bidi)."""
import io


def _org():
    import services as svc
    return svc.get_brand().get("org", "")


def _logo():
    import services as svc
    try:
        return svc.get_org_logo()   # (bytes, mime) or None
    except Exception:
        return None


def _shape(s):
    """Shape Arabic text (join letters, right-to-left) for correct PDF display."""
    s = "" if s is None else str(s)
    if any("\u0600" <= ch <= "\u06FF" or "\u0750" <= ch <= "\u077F" or
           "\uFB50" <= ch <= "\uFDFF" or "\uFE70" <= ch <= "\uFEFF" for ch in s):
        try:
            import arabic_reshaper
            from bidi.algorithm import get_display
            return get_display(arabic_reshaper.reshape(s))
        except Exception:
            return s
    return s


def xlsx_bytes(title, headers, rows, number_cols=None):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    org = _org()
    wb = Workbook(); ws = wb.active; ws.title = "Report"
    navy = PatternFill("solid", fgColor="1F3864"); bw = Font(bold=True, color="FFFFFF")
    ws["A1"] = org; ws["A1"].font = Font(bold=True, size=14, color="1F3864")
    ws["A2"] = title; ws["A2"].font = Font(bold=True, size=11, color="333333")
    r0 = 4
    for j, h in enumerate(headers, 1):
        c = ws.cell(row=r0, column=j, value=h); c.fill = navy; c.font = bw
        c.alignment = Alignment(horizontal="center")
    for i, row in enumerate(rows, start=r0 + 1):
        for j, v in enumerate(row, 1):
            ws.cell(row=i, column=j, value=v)   # Excel handles Arabic natively
    for j, h in enumerate(headers, 1):
        widest = max([len(str(h))] + [len(str(row[j - 1])) for row in rows]) if rows else len(str(h))
        ws.column_dimensions[get_column_letter(j)].width = min(42, max(10, widest + 2))
    ws.freeze_panes = f"A{r0 + 1}"
    buf = io.BytesIO(); wb.save(buf); return buf.getvalue()


def pdf_bytes(title, headers, rows):
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.units import mm
    from reportlab.lib import colors
    from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle, Paragraph,
                                    Spacer, Image, HRFlowable)
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    import arabicfont
    body_font = "Helvetica"
    try:
        body_font = arabicfont.register("Amiri")   # Arabic-capable
    except Exception:
        body_font = "Helvetica"

    NAVY = colors.HexColor("#1f3864")
    org = _org()
    wide = len(headers) > 7
    pagesize = landscape(A4) if wide else A4
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=pagesize, leftMargin=12 * mm, rightMargin=12 * mm,
                            topMargin=12 * mm, bottomMargin=12 * mm, title=title)
    org_s = ParagraphStyle("org", fontName="Helvetica-Bold", textColor=NAVY, fontSize=15, leading=17)
    title_s = ParagraphStyle("ttl", fontName="Helvetica", fontSize=11, textColor=colors.HexColor("#333333"), spaceAfter=6, spaceBefore=2)
    cell_s = ParagraphStyle("cell", fontName=body_font, fontSize=7, leading=9)
    head_s = ParagraphStyle("head", fontName="Helvetica-Bold", fontSize=7.5, leading=9, textColor=colors.white)
    story = []

    # ---- letterhead: small logo + org name on one row, title beneath ----
    logo_flowable = None
    logo = _logo()
    if logo:
        try:
            data, mime = logo
            if "svg" not in (mime or ""):
                img = Image(io.BytesIO(data))
                h = 15 * mm
                w = h * (img.imageWidth / float(img.imageHeight or 1))
                if w > 45 * mm:
                    w = 45 * mm; h = w * (img.imageHeight / float(img.imageWidth or 1))
                img.drawHeight = h; img.drawWidth = w
                logo_flowable = img
        except Exception:
            logo_flowable = None
    if logo_flowable is not None:
        head = Table([[logo_flowable, Paragraph(org, org_s)]], colWidths=[48 * mm, None], rowHeights=[16 * mm])
    else:
        head = Table([[Paragraph(org, org_s)]])
    head.setStyle(TableStyle([("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                              ("LEFTPADDING", (0, 0), (-1, -1), 0),
                              ("TOPPADDING", (0, 0), (-1, -1), 0),
                              ("BOTTOMPADDING", (0, 0), (-1, -1), 0)]))
    story.append(head)
    story.append(Paragraph(title, title_s))
    story.append(HRFlowable(width="100%", color=NAVY, thickness=1))
    story.append(Spacer(1, 6))

    # ---- table ----
    data = [[Paragraph(str(h), head_s) for h in headers]]
    for row in rows:
        data.append([Paragraph(_shape(v), cell_s) for v in row])
    tbl = Table(data, repeatRows=1)
    tbl.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), NAVY),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f3f5f9")]),
        ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#cccccc")),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 2), ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
        ("LEFTPADDING", (0, 0), (-1, -1), 3), ("RIGHTPADDING", (0, 0), (-1, -1), 3),
    ]))
    story.append(tbl)

    # ---- footer on every page: generated timestamp (left) + Page X of Y (right) ----
    from reportlab.pdfgen import canvas as _canvas
    from datetime import datetime as _dt
    stamp = "Generated " + _dt.now().strftime("%Y-%m-%d %H:%M") + "  \u00b7  " + org

    class _NumberedCanvas(_canvas.Canvas):
        def __init__(self, *a, **k):
            _canvas.Canvas.__init__(self, *a, **k)
            self._saved = []

        def showPage(self):
            self._saved.append(dict(self.__dict__))
            self._startPage()

        def save(self):
            total = len(self._saved)
            for state in self._saved:
                self.__dict__.update(state)
                self._footer(total)
                _canvas.Canvas.showPage(self)
            _canvas.Canvas.save(self)

        def _footer(self, total):
            w, h = self._pagesize
            self.setStrokeColor(colors.HexColor("#dddddd"))
            self.line(12 * mm, 11 * mm, w - 12 * mm, 11 * mm)
            self.setFont("Helvetica", 7)
            self.setFillColor(colors.HexColor("#777777"))
            self.drawString(12 * mm, 7 * mm, stamp)
            self.drawRightString(w - 12 * mm, 7 * mm, "Page %d of %d" % (self.getPageNumber(), total))

    doc.build(story, canvasmaker=_NumberedCanvas)
    return buf.getvalue()


XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
PDF_MIME = "application/pdf"
